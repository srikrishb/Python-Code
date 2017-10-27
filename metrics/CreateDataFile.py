import APIMethod as APIFile
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
import os

class CreateDataFile:

    def __init__(self, inputCriteria):
        self.inputCriteria = inputCriteria

    def postData(self, endpoint):

        fileEndpoint = 'output/' + endpoint
        payload = {'tableViewConfig': self.inputCriteria}

        response = APIFile.API.postSessionDataCall(fileEndpoint, payload)

        if response['statusCode'] == '1':
            return response['data']
        else:
            return 'Issue with parameters. Please check'

    # Creates the data file for specified parameters
    @staticmethod
    def createDataFile(targetMapList, dataFile):

        targetFileName = dataFile
        targetFileHeader = []
        targetFileRow = []
        targetFinalRowList = []
        mergeRow = {}
        i = 0
        if os.path.isfile(targetFileName):
            workbook = load_workbook(targetFileName)
        else:
            workbook = Workbook(targetFileRow)
            worksheet = workbook.active
            worksheet.title = 'Asset List'
            columnCount = 1
            # Write the File Header
            for innerMap in targetMapList:
                for key in innerMap:
                    map = innerMap[key]
                    if key == 'Asset Details':
                        if isinstance(map, dict):
                            for innerKey in map:
                                if innerKey not in targetFileHeader:
                                    targetFileHeader.append(innerKey)
                                    columnCount += 1
                        elif isinstance(map, list):
                            for innerList in map:
                                for innerKey in innerList:
                                    if innerKey not in targetFileHeader:
                                        targetFileHeader.append(innerKey)
                                        columnCount += 1
                    elif key == 'Attributes':
                        if 'Attribute Type' not in targetFileHeader:
                            targetFileHeader.append('Attribute Type')
                            targetFileHeader.append('Attribute')
                            columnCount += 2
                    elif key == 'Relations':
                        if 'Relation Type' not in targetFileHeader:
                            targetFileHeader.append('Relation Type')
                            targetFileHeader.append('Relation')
                            columnCount += 2
                    elif key == 'Complex Relations':
                        if 'Complex Relations - Relation Type' not in targetFileHeader:
                            targetFileHeader.append('Complex Relations - Name')
                            targetFileHeader.append('Complex Relations - Relation Type')
                            targetFileHeader.append('Complex Relations - Relation')
                            targetFileHeader.append('Complex Relations - Attribute Type')
                            targetFileHeader.append('Complex Relations - Attribute')
                            columnCount += 5
            worksheet.append(targetFileHeader)

            for columnIndex in range(1, columnCount):
                cell = worksheet.cell(row=1, column=columnIndex)
                cell.font = Font(bold=True)

        rowNum = 2
        col = 1
        # Write actual data
        for innerMap in targetMapList:

            prevHighestLen = 0
            for key in innerMap.keys():
                map = innerMap[key]
                if key == 'Asset Details':
                    if isinstance(map, dict):
                        for innerKey in map:
                            targetFileRow.append(map[innerKey])
                            worksheet.cell(row=rowNum, column=col).value = map[innerKey]
                            col += 1
                        highestLen = 1
                    elif isinstance(map, list):
                        for innerList in map:
                            for innerKey in innerList:
                                targetFileRow.append(innerList[innerKey])
                                worksheet.cell(row=rowNum, column=col).value = innerList[innerKey]
                                col += 1
                            highestLen = 1
                elif key == 'Attributes' or key == 'Relations':
                    highestLen = 0
                    currentRowNum = rowNum
                    for innerKey in map:
                        if isinstance(map[innerKey], list):
                            highestLen += len(map[innerKey])
                        else:
                            highestLen += 1
                        targetFileRow.append(innerKey)
                        targetFileRow.append(map[innerKey])
                        worksheet.cell(row=rowNum, column=col).value = innerKey[innerKey.find('>') + 1:]
                        col += 1
                        if isinstance(map[innerKey], list):
                            tempList = []
                            tempList = map[innerKey]
                            rowLen = len(tempList)
                            j = 0
                            for rowIndex in range(rowNum, rowNum + rowLen):
                                worksheet.cell(row=rowIndex, column=col).value = tempList[j]
                                j += 1
                            rowNum = rowNum + rowLen
                        else:
                            worksheet.cell(row=rowNum, column=col).value = map[innerKey]
                            rowNum = rowNum + 1
                        col -= 1
                    rowNum = currentRowNum
                    col += 2
                    if prevHighestLen > highestLen:
                        highestLen = prevHighestLen
                    prevHighestLen = highestLen

                elif key == 'Complex Relations':
                    currentRowNum = rowNum
                    currentColNum = col

                    for key in map:
                        highestLen = 0
                        innerMap = map[key]
                        if key == 'Relation':
                            for innerKey in innerMap:
                                col = currentColNum
                                # Name of Complex Relation Type
                                worksheet.cell(row=rowNum, column=col).value = innerKey
                                col += 1
                                innerInnerMap = innerMap[innerKey]
                                for innerInnerKey in innerInnerMap:
                                    # Name of the Relation Type under Complex Relation
                                    worksheet.cell(row=rowNum, column=col).value = innerInnerKey
                                    col += 1
                                    innerInnerInnerMap = innerInnerMap[innerInnerKey]
                                    # Relation Value
                                    if isinstance(innerInnerInnerMap, list):
                                        tempList = innerInnerInnerMap
                                        rowLen = len(tempList)
                                        j = 0
                                        for rowIndex in range(rowNum, rowNum + rowLen):
                                            worksheet.cell(row=rowIndex, column=col).value = tempList[j]
                                            j += 1
                                        rowNum = rowNum + rowLen
                                    else:
                                        worksheet.cell(row=rowNum, column=col).value = innerInnerInnerMap
                                        rowNum = rowNum + 1
                                    col -= 1

                        if key == 'Attributes':
                            rowNum = currentRowNum
                            for innerKey in innerMap:
                                # Name of Attribute Type
                                worksheet.cell(row=rowNum, column=col).value = innerKey
                                col += 1
                                innerInnerMap = innerMap[innerKey]
                                if isinstance(innerInnerMap, list):
                                    tempList = innerInnerMap
                                    rowLen = len(tempList)
                                    j = 0
                                    for rowIndex in range(rowNum, rowNum + rowLen):
                                        worksheet.cell(row=rowIndex, column=col).value = tempList[j]
                                        j += 1
                                    rowNum = rowNum + rowLen
                                else:
                                    worksheet.cell(row=rowNum, column=col).value = innerInnerMap
                                    rowNum = rowNum + 1
                                col -= 1

                        if prevHighestLen > highestLen:
                            highestLen = prevHighestLen
                        prevHighestLen = highestLen
                        rowNum = currentRowNum
                        col += 2

            targetFinalRowList.append(targetFileRow)
            mergeRow[rowNum] = highestLen
            rowNum = rowNum + highestLen
            targetFileRow = []
            col = 1

        col = 1
        for rowNum in mergeRow:
            for col in range(1, 5):
                worksheet.merge_cells(start_row=rowNum, start_column=col, end_row=rowNum + mergeRow[rowNum] - 1,
                                      end_column=col)
                col += 1

        worksheet.freeze_panes = 'A2'
        workbook.save(targetFileName)
        return targetFileName