from openpyxl import load_workbook
from openpyxl import Workbook
import redis
import os

class ProcessMetrics:

    # inputFileName is the data file that contains input data to kickoff analytics
    def __init__(self, inputFileName):
        self.inputFileName = inputFileName

    def generateMetricsFileI(self):
        # Open the data file
        workbook = load_workbook(self.inputFileName)
        defaultSheet = workbook.active

        # Kick start redis
        redis_db = redis.StrictRedis(host="127.0.0.1", port=6379, db=0)
        redis_db.flushall()
        i = 0
        key = ''
        offset = 2
        prev_assettype = ''
        curr_assettype = ''
        prev_offset = 2
        currAssetName = ''
        prevAssetName = ''
        keyList = []
        for row in defaultSheet.rows:
            if i == 0:
                i += 1
                for cellInRow in row:
                    if cellInRow.value == 'AssetType':
                        marker = cellInRow.col_idx
                        # xy = coordinate_from_string(cell)  # returns ('A',4)
                        # marker = column_index_from_string(xy[0])  # returns 1
                        print('marker', marker)
            else:
                for cellInRow in row:
                    if cellInRow.col_idx == 1:
                        currAssetName = cellInRow.value
                    if cellInRow.col_idx > 1 and cellInRow.col_idx < marker:
                        if cellInRow.col_idx == 2:
                            key = cellInRow.value
                        else:
                            key = key + ":" + cellInRow.value
                    if cellInRow.col_idx == marker:
                        key = key + ":" + cellInRow.value
                        curr_assettype = cellInRow.value
                        if prev_assettype == curr_assettype:
                            if prevAssetName == currAssetName:
                                offset = prev_offset
                            else:
                                offset += 1
                                prevAssetName = currAssetName
                                prev_offset = offset
                        else:
                            if prevAssetName != currAssetName:
                                offset = 2
                                prev_offset = offset
                                prev_assettype = curr_assettype
                                prevAssetName = currAssetName
                print(key, offset - 1, 1)
                keyList.append(key)

                # see what keys are in Redis

                redis_db.setbit(key, offset - 1, 1)
                key = ''

        for key in keyList:
            print(key, redis_db.bitcount(key))

    @staticmethod
    def checkLen(inputString):
        if str(inputString).strip() == '' or inputString is None :
            return 'No data available'
        if isinstance(inputString, bool):
            return str(inputString)
        else:
            return inputString

    def generateMetricsFileII(self, dimension, targetFileName):
        print('dimension', dimension)
        # Open the data file
        dataWorkbook = load_workbook(self.inputFileName)
        defaultSheet = dataWorkbook.active

        # Kick start redis
        redis_db = redis.StrictRedis(host="127.0.0.1", port=6379, db=0)
        redis_db.flushall()
        i = 0
        excludecolumnindex = 0
        marker = -1
        key = ''
        prevDimension = ''
        prevAssetName = ''
        keyList = {}
        targetFileHeader = []

        # Count columns in excel to determine whether any columns need to be excluded (such as asset name)
        for row in defaultSheet.rows:
            if i == 0:
                i += 1
                for cellInRow in row:
                    excludecolumnindex += 1
            else:
                break

        i = 0
        for row in defaultSheet.rows:

            if i == 0:
                i += 1
                for cellInRow in row:
                    if marker == -1 and cellInRow.col_idx > excludecolumnindex-1:
                        targetFileHeader.append(cellInRow.value)
                    if cellInRow.value == dimension:
                        marker = cellInRow.col_idx
            else:
                for cellInRow in row:

                    if cellInRow.col_idx == 1:
                        currAssetName = ProcessMetrics.checkLen(cellInRow.value)

                    if cellInRow.col_idx > 1 and cellInRow.col_idx < marker:
                        if cellInRow.col_idx == 2:
                            key = ProcessMetrics.checkLen(cellInRow.value)
                        else:
                            key = key + ":" + ProcessMetrics.checkLen(cellInRow.value)

                    if cellInRow.col_idx == marker:
                        if key == '':
                            key = ProcessMetrics.checkLen(cellInRow.value)
                        else:
                            key = key + ":" + ProcessMetrics.checkLen(cellInRow.value)

                        currDimension = ProcessMetrics.checkLen(cellInRow.value)
                        print('currDimension', currDimension, 'prevDimension', prevDimension)
                        # Compare the prev and curr dimension values
                        if prevDimension == currDimension:
                            print('prevAssetName', prevAssetName, 'currAssetName', currAssetName)
                            # If prev and curr dimension values don't match, compare the names of the assets to find unique assets
                            if prevAssetName != currAssetName:
                                # Increment score for the key
                                redis_db.zincrby('tempset', key, 1)
                                prevAssetName = currAssetName
                        else:
                            prevDimension = currDimension
                            print('prevDimension', prevDimension)
                            # If the prev and curr dimension values don't match, it implies that asset is new
                            if prevAssetName != currAssetName:
                                # Set the score for the new key
                                print('key', key)
                                redis_db.zadd('tempset', 1, key)
                                prevAssetName = currAssetName

                keyList[key] = ''
                key = ''

        if os.path.isfile(targetFileName):
            metricsWorkbook = load_workbook(targetFileName)
            metricsSheet = metricsWorkbook.create_sheet('Count by ' + dimension)
        else:
            metricsWorkbook = Workbook()
            metricsSheet = metricsWorkbook.active
            metricsSheet.title = 'Count by ' + dimension

        col = 1
        # Write Header
        for header in targetFileHeader:
            metricsSheet.cell(row=1, column=col).value = header
            col +=1

        metricsSheet.cell(row=1, column=col).value = 'Count'
        # Write metrics information
        row=2

        for key in keyList.keys():
            # Fetch the dimensions from key and print them in the file
            measure = redis_db.zscore('tempset', key)
            innerKeyList = key.split(":")

            # Initialize column index
            col = 1
            for innerKey in innerKeyList:
                metricsSheet.cell(row=row, column=col).value = innerKey
                # Increment column index to go to next cell in row
                col += 1

            # Fetch the measure value and print it in the file
            metricsSheet.cell(row=row, column=col).value = measure

            #Increment row counter to go to next line
            row += 1

        # Persist the data in excel file
        metricsWorkbook.save(targetFileName)