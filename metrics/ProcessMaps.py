import re
from AttributeFilter import AttributeFilter
from RelationFilter import RelationFilter
from ComplexRelationFilter import ComplexRelationFilter
from Asset import Asset
from CreateMap import CreateMap
from CreateDataFile import CreateDataFile
import os
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
import time
import math

class ProcessMaps:

    outputParameterList = []

    def __init__(self, inputMap):
        self.inputMap = inputMap

    @staticmethod
    def cleanhtml(raw_html):
      cleanr = re.compile('<.*?>')
      cleantext = re.sub(cleanr, '', raw_html)
      cleantext = cleantext.replace(u'\xa0', u' ')
      return cleantext

    def processMaps(self, detailedFile):

        targetData = ''
        outputFilter = ''
        fileNameList = []
        finalMap = {}
        finalResultList = []
        eachMap = self.inputMap
        for itemkey in eachMap.keys():
            if itemkey == 'conditions':
                outputFilter = eachMap[itemkey].keys()
                if 'relationFilter' not in outputFilter and 'complexRelationFilter' not in outputFilter and 'attributeFilter' not in outputFilter:
                    if isinstance(eachMap[itemkey], dict):
                        targetData = CreateMap.generateMap(eachMap[itemkey])
                elif 'attributeFilter' in outputFilter:
                    tempMap = {}
                    for valuesMap in eachMap[itemkey].values():
                        for innerKey in valuesMap:
                            tempMap[innerKey] = valuesMap[innerKey]
                            if innerKey == 'attributeTypeEquals':
                                assetFilterObj = AttributeFilter(tempMap)
                                attributeKey = valuesMap[innerKey]
                                tempTargetData = Asset.fetchDataSet(innerKey, '')
                                preTargetData = assetFilterObj.filterAttributeDataSet(tempTargetData)
                                targetData = preTargetData
                                tempMap = {}
                            elif innerKey == 'attributeValueEquals':
                                filterAssetTypes = AttributeFilter(tempMap)
                                targetData = filterAssetTypes.filterTargetDataSet(attributeKey, preTargetData)
                elif 'complexRelationFilter' in outputFilter:
                    tempMap = {}
                    for valuesMap in eachMap[itemkey].values():
                        for innerKey in valuesMap:
                            tempMap[innerKey] = valuesMap[innerKey]
                            if innerKey == 'complexRelationName':
                                complexRelationFilterObj = ComplexRelationFilter(tempMap)
                                tempTargetData = Asset.fetchDataSet(innerKey,'')
                                preTargetData = complexRelationFilterObj.filterComplexRelationDataSet(tempTargetData)
                                targetData = preTargetData
                                tempMap = {}
                            elif innerKey == 'AssetType':
                                filterAssetTypes = ComplexRelationFilter(tempMap)
                                targetData = filterAssetTypes.filterTargetDataSet(preTargetData)
                elif 'relationFilter' in outputFilter:
                    tempMap = {}
                    for valuesMap in eachMap[itemkey].values():
                        for innerKey in valuesMap:
                            tempMap[innerKey] = valuesMap[innerKey]
                            if innerKey == 'RelationTypeIn':
                                relationFilterObj = RelationFilter(tempMap)
                                tempTargetData = Asset.fetchDataSet(innerKey, '')
                                preTargetData = relationFilterObj.filterRelationDataSet(tempTargetData)
                                targetData = preTargetData
                                tempMap = {}
                            elif innerKey == 'AssetType':
                                filterAssetTypes = RelationFilter(tempMap)
                                targetData = filterAssetTypes.filterTargetDataSet(preTargetData)
            elif itemkey == 'outputResult':
                if outputFilter != 'relationPath':
                    outputResultParameters = eachMap[itemkey]
                    for i in range(0, len(targetData)):
                        targetDataMap = {}
                        assetObj = Asset(targetData[i]['resourceId'])
                        for outputParameter in outputResultParameters.keys():

                            # targetDataMap['Status'] = targetData[i]['statusReference']['signifier']
                            if outputParameter == 'Asset Details':
                                ProcessMaps.outputParameterList = outputResultParameters[outputParameter]
                                for detailType in ProcessMaps.outputParameterList:
                                    targetDataMap[detailType] = assetObj.fetchAssetAssetDetails(detailType, targetData[i])

                        tempMap = {}
                        finalMap['Asset Details'] = targetDataMap
                        finalResultList.append(finalMap)
                        finalMap = {}

                    # Create the output file
                    fileName = CreateDataFile.createDataFile(finalResultList, detailedFile)
                    if fileName not in fileNameList:
                       fileNameList.append(fileName)

        return targetData

    @staticmethod
    def checkCompletion(inputAssetData, detailedFile):
        finalMap = {}
        finalResultList = []
        # Fetch Completion Status for all the assets
        for i in range(0, len(inputAssetData)):

            targetDataMap = {}
            assetObj = Asset(inputAssetData[i]['resourceId'])
            for detailType in ProcessMaps.outputParameterList:
                targetDataMap[detailType] = assetObj.fetchAssetAssetDetails(detailType, inputAssetData[i])

            # Find possible attributes types and relations types
            possibleRelationsAndAttributesResponse = assetObj.fetchPossibleRelationsTypesAndAttributesTypes()
            possibleAttributeTypesList = []
            possibleRelationTypesList = []

            if possibleRelationsAndAttributesResponse != 'No Data Found':
                possibleRelationsAndAttributesList = possibleRelationsAndAttributesResponse['representationReference']
                for k in range(0, len(possibleRelationsAndAttributesList)):
                    if 'descriptionReference' in possibleRelationsAndAttributesList[k].keys():
                        possibleAttributeTypesList.append(possibleRelationsAndAttributesList[k]['signifier'])

                    if 'role' in possibleRelationsAndAttributesList[k].keys():
                        if possibleRelationsAndAttributesList[k]['role'] not in possibleRelationTypesList:
                            possibleRelationTypesList.append(possibleRelationsAndAttributesList[k]['role'])
                        if possibleRelationsAndAttributesList[k]['coRole'] not in possibleRelationTypesList:
                            possibleRelationTypesList.append(possibleRelationsAndAttributesList[k]['coRole'])
            completion = 'Y'
            # Fetch relation types that have been defined for an asset
            targetDataRelationsList = []
            relationsResponse = assetObj.fetchRelations()
            if relationsResponse != 'No Data Found':
                # Add all the relation types that have a value
                for relationResponseMap in relationsResponse:
                    for innerRelationsMap in relationResponseMap['relation']:
                        if 'vocabularyReference' in innerRelationsMap['typeReference']['headTerm']:
                            relationType = innerRelationsMap['typeReference']['headTerm']['vocabularyReference']['name']

                        if relationType != 'Complex Relation Types':
                            targetDataRelationsList.append(innerRelationsMap['typeReference']['role'])
                            targetDataRelationsList.append(innerRelationsMap['typeReference']['coRole'])

                # Determine whether all relation types have been defined
                for possibleRelationType in possibleRelationTypesList:
                        if possibleRelationType not in targetDataRelationsList:
                            targetDataMap['Completion-Category'] = 'N'
                            completion = 'N'
                            break

            # Fetch attributes types that have been defined for an asset
            if completion == 'Y':
                targetDataAttributesList = []
                attributesResponse = assetObj.fetchAttributes()
                if attributesResponse != 'No Data Found':
                    attributesResponseList = attributesResponse['attributeReference']

                    for listIndex in range(0, len(attributesResponseList)):
                        attributeResponseMap = attributesResponseList[listIndex]
                        key = attributeResponseMap['labelReference']['signifier']
                        if key not in targetDataAttributesList:
                            targetDataAttributesList.append(key)

                    # Add all the attribute types that don't have a value
                    for possibleAttribute in possibleAttributeTypesList:
                        if possibleAttribute not in targetDataAttributesList:
                            targetDataMap['Completion-Category'] = 'N'
                            completion = 'N'
                            break

            if 'Completion-Category' not in targetDataMap.keys():
                targetDataMap['Completion-Category'] = 'Y'

            finalMap['Asset Details'] = targetDataMap
            finalResultList.append(finalMap)
            finalMap = {}


        #Print the results to a file
        fileName = CreateDataFile.createDataFile(finalResultList, detailedFile)
        return fileName

    @staticmethod
    def checkProcess(inputAssetData, detailedFile, metricsFile):
        for i in range(0, len(inputAssetData)):
            targetDataMap = {}
            tempMap = {}
            assetObj = Asset(inputAssetData[i]['resourceId'])
            maxCol = 0

            for detailType in ProcessMaps.outputParameterList:
                targetDataMap[detailType] = assetObj.fetchAssetAssetDetails(detailType, inputAssetData[i])
                tempMap[detailType] = targetDataMap[detailType]

            # Find all processes for the asset
            processResponse = assetObj.fetchProcesses()
            if processResponse != 'No Data Found':
                if len(processResponse) > 1:
                    valueList = []
                    for process in processResponse:
                        targetDataMap['Process-Category'] = process
                        valueList.append(targetDataMap)

                        targetFileName = detailedFile
                        targetFileHeader = []
                        targetFileRow = []
                        targetFinalRowList = []
                        i = 0
                        if os.path.isfile(targetFileName):
                            workbook = load_workbook(targetFileName)
                            worksheet = workbook.get_sheet_by_name('Asset List')
                        else:
                            workbook = Workbook(targetFileRow)
                            worksheet = workbook.active
                            worksheet.title = 'Asset List'
                            columnCount = 1
                            # Write the File Header
                            for innerMap in valueList:
                                for key in innerMap:
                                    if key not in targetFileHeader:
                                        targetFileHeader.append(key)
                                        columnCount += 1

                            worksheet.append(targetFileHeader)

                            for columnIndex in range(1, columnCount):
                                cell = worksheet.cell(row=1, column=columnIndex)
                                cell.font = Font(bold=True)

                            rowNum = 2
                            col = 1

                        # Write actual data
                        for innerMap in valueList:
                            for key in innerMap.keys():
                                targetFileRow.append(innerMap[key])
                                worksheet.cell(row=rowNum, column=col).value =innerMap[key]
                                workbook.save(targetFileName)

                                if col > maxCol:
                                    maxCol = col
                                col += 1

                            targetFinalRowList.append(targetFileRow)
                            rowNum = rowNum + 1
                            targetFileRow = []
                            col = 1

                        valueList = []

        workbook = load_workbook(targetFileName)
        worksheet = workbook.get_sheet_by_name('Asset List')

        assetList = []
        outputMap = {}
        for row in range(2, rowNum):
            for col in (maxCol, 1, -1):
                if col == maxCol:
                    key = worksheet.cell(row=row, column=col).value
                    if key in outputMap.keys():
                        assetList = outputMap[key]
                    else:
                        assetList = []
                if col == 1:
                    assetList.append(worksheet.cell(row=row, column=col).value)
                outputMap[key] = assetList

        metricsWorkbook = Workbook()
        metricsWorksheet = metricsWorkbook.active
        metricsWorksheet.title = 'Counts by Process-Category'

        metricsWorksheet.cell(row=1, column=1).value = 'Process-Category'
        metricsWorksheet.cell(row=1, column=2).value = 'Counts'
        metricsWorksheet.cell(row=1, column=1).font = Font(bold=True)
        metricsWorksheet.cell(row=1, column=2).font = Font(bold=True)

        row = 2
        col = 1
        for key in outputMap.keys():
            metricsWorksheet.cell(row = row, column = col).value = key
            col += 1
            metricsWorksheet.cell(row=row, column=col).value = len(outputMap[key])
            col -= 1
            row += 1

        metricsWorkbook.save(metricsFile)
        return targetFileName

    def fetchLastViewedAssets(self, detailedFile):
        lastViewedAssetsEndpoint = 'navigation/most_viewed'
        inputDate = int(time.mktime(time.strptime(self.inputMap, '%Y-%m-%d %H:%M:%S')))
        lastViewedAssetsPayload = {'timerange': math.ceil(time.time() - inputDate)*1000}
        lastViewedAssetsResponse = Asset.getDataCall(lastViewedAssetsEndpoint,lastViewedAssetsPayload)

        print(math.ceil(time.time() - inputDate))
        if lastViewedAssetsResponse['statusCode'] != '1':
            return 'No Data Found'
        else:
            lastViewedAssetsResults = lastViewedAssetsResponse['data']
            if len(lastViewedAssetsResults) == 0:
                return 'No Data Fetched'
            else:
                # Save data in workbook
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = 'Asset List'
                rowNum = 1
                colNum = 1

                # Write the header
                targetFileHeader = ['Community', 'Domain', 'Asset Name', 'Asset Type', 'Number of Views', 'Last Viewed Date']
                for targetHeader in targetFileHeader:
                    worksheet.cell(row=rowNum, column=colNum).value = targetHeader
                    cell = worksheet.cell(row=rowNum, column=colNum)
                    cell.font = Font(bold=True)
                    colNum += 1

                targetFileName = detailedFile

                # Read data from result and write the data
                rowNum +=1
                colNum = 1

                for assetMap in lastViewedAssetsResults:
                    assetResponse = Asset.getDataCall('term/'+assetMap['termId'], '')

                    if assetResponse['statusCode'] != '1':
                        print('No data received')
                    else:
                        assetDetails = assetResponse['data']
                        # Community Namme
                        worksheet.cell(row=rowNum, column=colNum).value = Asset.fetchAssetAssetDetails('Community', assetDetails)
                        colNum += 1

                        # Domain Name
                        worksheet.cell(row=rowNum, column=colNum).value = Asset.fetchAssetAssetDetails('Domain', assetDetails)
                        colNum += 1

                        # Asset Name
                        worksheet.cell(row=rowNum, column=colNum).value = assetMap['signifier']
                        colNum += 1

                        # Asset Type
                        worksheet.cell(row=rowNum, column=colNum).value = Asset.fetchAssetAssetDetails('Asset Type', assetDetails)
                        colNum += 1

                        # Number of views
                        worksheet.cell(row=rowNum, column=colNum).value = assetMap['nbViews']
                        colNum += 1

                        # Last Viewed Date
                        worksheet.cell(row=rowNum, column=colNum).value = Asset.convertEpochTime(assetMap['lastViewedDate'])
                        rowNum += 1
                        colNum = 1

                workbook.save(targetFileName)

                return 'File saved successfully'



