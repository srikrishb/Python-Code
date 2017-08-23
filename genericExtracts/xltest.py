from openpyxl import load_workbook

filename = 'K:\Git Code\Python\Output\file.xlsx'
wb = load_workbook(filename)
ws1 = wb.create_sheet("test")

temp = [['Amortized Cost (USD Equivalent)'], '01. Amortized Cost (USD Equivalent)', ['Accretion of discount', 'Adjusted purchase price', 'Face value', 'Purchase price']]

mergeLen = 0
totalLen = len(temp)
i = 0
for col in range(1,totalLen+1):
    if isinstance(temp[i], str):
        ws1.cell(row=1, column=col).value = temp[i]
        i += 1
    elif isinstance(temp[i], list):
        tempList = []
        tempList = temp[i]
        rowLen = len(tempList)
        if mergeLen < rowLen:
            mergeLen = rowLen
        j = 0
        for rowIndex in range(1, rowLen+1):
            ws1.cell(row=rowIndex, column=col).value = tempList[j]
            j += 1
        i += 1

i = 0
for col in range(1,totalLen+1):
    if (isinstance(temp[i], list) and len(temp[i]) == 1) or isinstance(temp[i], str):
        ws1.merge_cells(start_row=1, start_column=col, end_row=mergeLen, end_column=col)

    elif isinstance(temp[i], list) and len(temp[i]) > 1:
        print('No Prob')
    i += 1

ws2 = wb.get_sheet_names()

print(ws2)


wb.save(filename=filename)

