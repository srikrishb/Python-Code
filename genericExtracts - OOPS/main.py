#!/usr/bin/ python

import os
import sys
import yaml
import time
import APIMethod as APIFile
import csv
import math
import smtplib
import shutil
import mimetypes
from email import encoders
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.audio import MIMEAudio
from email.mime.base import MIMEBase
from email.utils import COMMASPACE, formatdate
import re
from zipfile import ZipFile
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color

def cleanhtml(raw_html):
  cleanr = re.compile('<.*?>')
  cleantext = re.sub(cleanr, '', raw_html)
  cleantext = cleantext.replace(u'\xa0', u' ')
  return cleantext

def getDataCall(endpoint, payload):
    # Kickoff the workflow using the endpoint and payload
    trylogin = APIFile.API.getCall(endpoint, payload)

    return trylogin

def fetchDataSet(innerMapKey, innerMapValue):

    result = dict()
    if str.upper(innerMapKey).find("LASTMODIFIED") != -1 or str.upper(innerMapKey).find("CREATED") != -1 or str.upper(innerMapKey).find('ASSETNAMEIN') != -1 or str.upper(innerMapKey).find('ATTRIBUTETYPE') != -1:
        searchSignfier = ''
    elif str.upper(innerMapKey) == 'ASSETNAMELIKE' or str.upper(innerMapKey) == 'ASSETNAMEEQUALS' :
        searchSignfier = innerMapValue


    if str.upper(innerMapKey).find('ATTRIBUTETYPE') != -1 or str.upper(innerMapKey) == 'ASSETNAMEIN' or str.upper(innerMapKey) == 'ASSETNAMELIKE' or str.upper(innerMapKey) == 'ASSETNAMEEQUALS' or str.upper(innerMapKey).find("LASTMODIFIED") != -1 or str.upper(innerMapKey).find("CREATED") != -1:
        assetNameLikeEndpoint = 'term/find/full'
        assetNameLikePayload = {'excludeMeta': 'false', 'searchSignifier': searchSignfier}
        assetNameLikeResponse = getDataCall(assetNameLikeEndpoint, assetNameLikePayload)

        if assetNameLikeResponse['statusCode'] == '1':
            result = assetNameLikeResponse['data']['term']

    return result

def convertEpochTime(inputTime):

    mathTargetTime = math.ceil(int(inputTime) / 1000)
    return time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(mathTargetTime))

def filterAssetData(filterAttributeType, innerMap, dataTobeFiltered):
    filteredData = []
    for key in innerMap.keys():
        filterValue = innerMap[key]
        if str.upper(key).find("ATTRIBUTEVALUEEQUALS") != -1:
            for data in dataTobeFiltered:
                # Find the resourceId
                attributesResponse = fetchAttributes(data['resourceId'])
                attributesResponseList = attributesResponse['attributeReference']

                existingValue = []
                # Loop through the attributesResponseList to check whether specified filterValue has been set
                for listIndex in range(0, len(attributesResponseList)):
                    attributeResponseMap = attributesResponseList[listIndex]
                    # key = attributeResponseMap['resourceId'] + ' -> ' + attributeResponseMap['labelReference']['signifier']
                    attributeType = attributeResponseMap['labelReference']['signifier']
                    if attributeType == filterAttributeType:
                        cleanedValue = cleanhtml(attributeResponseMap['value'])
                        if cleanedValue == filterValue:
                            filteredData.append(data)

    return filteredData

def fetchPossibleAttributes(resourceId):
    possibleRandAEndpoint = 'concept_type/' + resourceId + '/possible_attribute_types'
    possibleRandAPayload = ''
    possibleRandAResponse = getDataCall(possibleRandAEndpoint, possibleRandAPayload)
    if possibleRandAResponse['statusCode'] == '1':
        return possibleRandAResponse['data']
    else:
        return 'No Data Found'

def filterTargetData(innerMap, dataToBeFiltered):
    filteredData = []
    for key in innerMap.keys():

        filterValue = innerMap[key]

        if str.upper(key).find("LASTMODIFIED") != -1 or str.upper(key).find("CREATED") != -1:
            if str.upper(key).find("LASTMODIFIED") != -1:
                lookupKey = 'lastModified'
            elif str.upper(key).find("CREATEDBEFORE") or key.find("CREATEDAFTER") or key.find("CREATEDBETWEEN") != -1:
                lookupKey = 'createdOn'

            i = 0
            for data in dataToBeFiltered:

                dateTargetTime = convertEpochTime(dataToBeFiltered[i][lookupKey])
                if str.upper(key).find("AFTER") !=-1:
                    if dateTargetTime > filterValue:
                        filteredData.append(data)
                elif str.upper(key).find("BEFORE") !=-1:
                    if dateTargetTime < filterValue:
                        filteredData.append(data)
                elif str.upper(key).find("BETWEEN") !=-1:

                    term1 = filterValue[0]
                    term2 = filterValue[1]
                    if dateTargetTime >= term1 and dateTargetTime <= term2:
                        filteredData.append(data)

                i += 1
        elif str.upper(key).find("ASSETNAMELIKE") != -1:
            return dataToBeFiltered

        elif str.upper(key).find("ASSETNAMEEQUALS") != -1:
            i = 0
            for data in dataToBeFiltered:
                if data['signifier'] == filterValue:
                    filteredData.append(data)
                i += 1

        elif str.upper(key).find("ASSETNAMEIN") != -1:
            i = 0
            for filterItem in filterValue:
                for data in dataToBeFiltered:
                    if data['signifier'] == filterItem:
                        filteredData.append(data)
                    i += 1

        elif str.upper(key).find("ATTRIBUTETYPE") != -1:
            i = 0
            for data in dataToBeFiltered:
                #Find out whether there are any attributes tied to the asset:
                possibleAttributesList = fetchPossibleAttributes(data['resourceId'])
                if len(possibleAttributesList['attributeType']) > 0:
                    #Fetch attributes for the resourceId
                    attributesResponse = fetchAttributes(data['resourceId'])
                    attributesResponseList = attributesResponse['attributeReference']

                    #Loop through the attributesResponseList to check whether specified filterValue has been set
                    for listIndex in range(0, len(attributesResponseList)):
                        attributeResponseMap = attributesResponseList[listIndex]
                        attributeType = attributeResponseMap['labelReference']['signifier']
                        if attributeType == filterValue:
                            filteredData.append(data)
                    i += 1

    return filteredData

def listUnion(refList , dataList):

    newList = []
    newMap = refList

    for item in refList:
        newList.append(item['resourceId'])

    for items2 in dataList:
        if items2['resourceId'] not in newList:
            newMap.append(items2)

    return newMap

def listIntersect(refList , dataList):

    return set(refList).intersection(dataList)

def createDataSet(innerMap):
    tempMap = {}
    targetData = {}
    for key in innerMap.keys():
        # if isinstance(innerMap[key], dict) == 'false':
        if key == 'Or' or key == 'And':
            criteriaMap = innerMap[key]
            i = 0
            for criteriaKey in criteriaMap.keys():
                tempMap[criteriaKey] = criteriaMap[criteriaKey]
                if i == 0:
                    i += 1
                    tempTargetData = fetchDataSet(criteriaKey, criteriaMap[criteriaKey])
                    preTargetData = filterTargetData(tempMap, tempTargetData)
                else:
                    if key == 'And':
                        targetData = filterTargetData(tempMap, preTargetData)
                        preTargetData = targetData
                    if key == 'Or':
                        tempTargetData = fetchDataSet(criteriaKey, criteriaMap[criteriaKey])
                        newTargetData = filterTargetData(tempMap, tempTargetData)
                        targetData = listUnion(newTargetData, preTargetData)
                tempMap = {}
    return targetData

def checkNestedMaps(checkMap):
    keyList = []
    for innerMap in checkMap:
        keyList.append(innerMap)
    if 'Or' in keyList or 'And' in keyList:
        return 1
    else:
        return 0

def createMap(innerMap):
    tempMap = {}
    targetData = {}
    testMap = {}
    parameterMap = {}
    for key in innerMap.keys():
        # if isinstance(innerMap[key], dict) == 'false':
        if key == 'Or' or key == 'And':
            criteriaMap = innerMap[key]
            testMap[key] = innerMap[key]
            if checkNestedMaps(criteriaMap) == 1:
                preTargetData = createMap(criteriaMap)

            else:
                i = 0
                for testKey in testMap.keys():
                    for criteriaKey in testMap[testKey]:
                        parameterMap = testMap[testKey]
                        tempMap[criteriaKey] = parameterMap[criteriaKey]
                        if i == 0:
                            i += 1
                            tempTargetData = fetchDataSet(criteriaKey, parameterMap[criteriaKey])
                            preTargetData = filterTargetData(tempMap, tempTargetData)
                        else:
                            if testKey == 'And':
                                targetData = filterTargetData(tempMap, preTargetData)
                                preTargetData = targetData
                            if testKey == 'Or':
                                tempTargetData = fetchDataSet(criteriaKey, parameterMap[criteriaKey])
                                newTargetData = filterTargetData(tempMap, tempTargetData)
                                targetData = listUnion(newTargetData, preTargetData)
                        tempMap = {}



    return targetData

def fetchAttributeFilterDataSet(inputMap):
    i = 0
    tempMap = {}
    for inputKey in inputMap:
        innerList = inputMap[inputKey]
        for innerMap in innerList:
            for innerKey in innerMap:
                tempMap[innerKey] = innerMap[innerKey]
                if i == 0:
                    i += 1
                    tempTargetData = fetchDataSet(innerKey, '')
                    attributeKey = innerMap[innerKey]
                    preTargetData = filterTargetData(tempMap, tempTargetData)
                else:
                    targetData = filterAssetData(attributeKey, tempMap, preTargetData)
                tempMap = {}

    return targetData

def fetchRelationFilterDataSet(inputMap):
    i = 0
    tempMap = {}
    for inputKey in inputMap:
        innerList = inputMap[inputKey]
        if len(innerList) > 1:
            primaryAssetType = innerList[0]

        else:
            print('Need more than 1 assettype in List')

def generateFile(inputMap):
    i = 0
    tempMap = {}

    for masterKey in inputMap:
        toBeProcessedMap = inputMap[masterKey]
        print(toBeProcessedMap, type(toBeProcessedMap))
        if isinstance(toBeProcessedMap, dict):
            for innerMapKey in toBeProcessedMap:
                    if masterKey == 'Or' or masterKey == 'And':
                        tempMap[innerMapKey] = toBeProcessedMap[innerMapKey]
                        if innerMapKey == 'attributeFilter':
                            tempTargetData = fetchAttributeFilterDataSet(tempMap)
                        else:
                            tempTargetData = fetchDataSet(innerMapKey, toBeProcessedMap[innerMapKey])

                        if i == 0:
                            i += 1
                            preTargetData = filterTargetData(tempMap, tempTargetData)
                        else:

                            if innerMapKey == 'attributeFilter':
                                newTargetData = tempTargetData
                            else:
                                newTargetData = filterTargetData(tempMap, tempTargetData)

                            if masterKey == 'And':
                                targetData = listIntersect(newTargetData, preTargetData)

                            if masterKey == 'Or':
                                targetData = listUnion(newTargetData, preTargetData)

                            preTargetData = targetData

                        tempMap = {}

        elif masterKey == 'attributeFilter':
            tempMap[masterKey] = inputMap[masterKey]
            targetData = fetchAttributeFilterDataSet(tempMap)
            tempMap = {}

        elif masterKey == 'relationPath':
            tempMap[masterKey] = inputMap[masterKey]
            targetData = fetchRelationPathDataSet(tempMap)
            tempMap = {}
        elif isinstance(toBeProcessedMap, str) or isinstance(toBeProcessedMap, list):
            tempTargetData = fetchDataSet(masterKey, inputMap[masterKey])
            targetData = filterTargetData(inputMap, tempTargetData)

    return targetData

def createMapII(inputMap):
    tempMap = {}
    for key in inputMap:
        if checkNestedMaps(inputMap[key]) == 0:
            tempMap[key] = inputMap[key]
            targetData = generateFile(tempMap)
            tempMap = {}

    return targetData

def fetchAssetAssetDetails(detailType, targetData):

    if detailType == 'Asset Name':
        return targetData['signifier']

    if detailType == 'Community':
        if 'vocabularyReference' in targetData:
            if 'communityReference' in targetData['vocabularyReference']:
                if 'name' in targetData['vocabularyReference']['communityReference']:
                    return targetData['vocabularyReference']['communityReference']['name']
                else:
                    return 'Not Applicable'
            else:
                return 'Not Applicable'
        else:
            return 'Not Applicable'

    if detailType == 'Domain':
        if 'vocabularyReference' in targetData:
            if 'name' in targetData['vocabularyReference']:
                return targetData['vocabularyReference']['name']
            else:
                return 'Not Applicable'
        else:
            return 'Not Applicable'

    if detailType == 'Asset Type':
        if 'conceptType' in targetData:
            if 'signifier' in targetData['conceptType']:
                return targetData['conceptType']['signifier']
            else:
                return 'Not Applicable'
        else:
            return'Not Applicable'

    if detailType == 'Status':
        if 'statusReference' in targetData:
            if 'signifier' in targetData['statusReference']:
                #targetDataMap['Status'] = targetData[i]['statusReference']['signifier']
                return targetData[i]['statusReference']['signifier']
            else:
                return 'Not Applicable'
        else:
            return 'Not Applicable'

    if detailType == 'Articulation Score':
        return targetData.get('articulation', 'Not Applicable')

    if detailType == 'Created On':
        return convertEpochTime(targetData.get('createdOn'))
    else:
        return 'Not Applicable'

    if detailType == 'Created By':
        return targetData['createdBy']['firstName'] + targetData['createdBy']['lastName']
    else:
        return 'Not Applicable'

    if detailType == 'Last Modified':
        return convertEpochTime(targetData.get('lastModified'))
    else:
        return 'Not Applicable'

    if detailType == 'Last Modified By':
        return targetData['lastModifiedBy']['firstName'] + targetData['lastModifiedBy']['lastName']
    else:
        return 'Not Applicable'

def fetchPossibleRelationsAndAttributes(resourceId):
    possibleRandAEndpoint = 'concept_type/' + resourceId + '/possible_attribute_relation_types'
    possibleRandAPayload = ''
    possibleRandAResponse = getDataCall(possibleRandAEndpoint, possibleRandAPayload)
    if possibleRandAResponse['statusCode'] == '1':
        return possibleRandAResponse['data']
    else:
        return 'No Data Found'

def fetchRelations(resourceId):
    relationsResponseList = []

    sourceError = 'true'
    targetError = 'true'

    sourceRelationsEndpoint = 'concept_type/' + resourceId + '/source_relations'
    sourceRelationsPayload = ''
    sourceRelationsResponse = getDataCall(sourceRelationsEndpoint, sourceRelationsPayload)
    if sourceRelationsResponse['statusCode'] == '1':
        relationsResponseList.append(sourceRelationsResponse['data'])
    else:
        sourceError = 'No Data Found'

    targetRelationsEndpoint = 'concept_type/' + resourceId + '/target_relations'
    targetRelationsPayload = ''
    targetRelationsResponse = getDataCall(targetRelationsEndpoint, targetRelationsPayload)
    if targetRelationsResponse['statusCode'] == '1':
        relationsResponseList.append(targetRelationsResponse['data'])
    else:
        targetError = 'No Data Found'

    if sourceError == 'true' and targetError == 'true':
        return relationsResponseList
    else:
        return 'No Data Found'

def fetchAttributes(resourceId):

    attributesEndpoint = 'concept_type/' + resourceId + '/attributes'
    attributesPayload = ''
    attributesResponse = getDataCall(attributesEndpoint, attributesPayload)

    if attributesResponse['statusCode'] == '1':
        return attributesResponse['data']

def fetchComplexRelations(resourceId):
    complexRelationDefinitionEndpoint = 'complex_relation/'
    complexRelationDefinitionPayload = {'term': resourceId}
    complexRelationDefinitionResponse = getDataCall(complexRelationDefinitionEndpoint, complexRelationDefinitionPayload)

    if complexRelationDefinitionResponse['statusCode'] == '1':
        complexRelationDefinitionData = complexRelationDefinitionResponse['data']
        if len(complexRelationDefinitionData['termReference']) == 0:
            return 'No Relations Found'
        else:
            complexRelationRelationsEndpoint = 'complex_relation/' + complexRelationDefinitionData['termReference'][0]['resourceId'] + '/relations'
            complexRelationRelationsPayload = ''
            complexRelationRelationsResponse = getDataCall(complexRelationRelationsEndpoint, complexRelationRelationsPayload)
            if complexRelationRelationsResponse['statusCode'] == '1':
                return complexRelationRelationsResponse['data']
            else:
                return 'No Data Found'

def fetchComplexRelationsAttributes(resourceId):
    complexRelationDefinitionEndpoint = 'complex_relation/'
    complexRelationDefinitionPayload = {'term': resourceId}
    complexRelationDefinitionResponse = getDataCall(complexRelationDefinitionEndpoint,
                                                    complexRelationDefinitionPayload)

    if complexRelationDefinitionResponse['statusCode'] == '1':
        complexRelationDefinitionData = complexRelationDefinitionResponse['data']
        if len(complexRelationDefinitionData['termReference']) == 0:
            return 'No Attributes Found'
        else:
            complexRelationAttributesEndpoint = 'complex_relation/' + complexRelationDefinitionData['termReference'][0]['resourceId']
            complexRelationAttributesPayload = ''
            complexRelationAttributesResponse = getDataCall(complexRelationAttributesEndpoint,complexRelationAttributesPayload)
            if complexRelationAttributesResponse['statusCode'] == '1':
                return complexRelationAttributesResponse['data']
            else:
                return 'No Data Found'

def createTargetDataFile(targetMap, fileNamePrefix, fileNameSuffix):

    targetFileName = 'K:/Git Code/Python/Output/' + fileNamePrefix + fileNameSuffix
    with open(targetFileName, 'a', newline='') as targetFile:
        csvWriter = csv.writer(targetFile, delimiter=',')

        #{'Name': {'signifier': 'Amortized Cost (USD Equivalent)'}}

        # Write the headers of the asset into the file
        # targetFileRow = []
        # for mainHeaderKey in targetMap.keys():
        #     subHeaderMap = targetMap[mainHeaderKey]
        #     for subHeaderKey in subHeaderMap:
        #         targetFileRow.append(mainHeaderKey)
        # csvWriter.writerow(targetFileRow)

        # Write the sub-headers of the asset into the file
        targetFileRow = []
        for mainHeaderKey in targetMap.keys():
            subHeaderMap = targetMap[mainHeaderKey]
            for subHeaderKey in subHeaderMap.keys():
                if '->' in subHeaderKey:
                    targetFileRow.append(subHeaderKey[subHeaderKey.index('->')+3:])
                else:
                    targetFileRow.append(subHeaderKey)

                if isinstance(subHeaderMap[subHeaderKey], list):
                    for listItem in subHeaderMap[subHeaderKey]:
                        targetFileRow.append(listItem)
                else:
                    targetFileRow.append(subHeaderMap[subHeaderKey])

                csvWriter.writerow(targetFileRow)
                targetFileRow = []

        # Write the actual values of the asset into the file
        # targetFileRow = []
        # for mainHeaderKey in targetMap.keys():
        #     subHeaderMap = targetMap[mainHeaderKey]
        #     for subHeaderKey in subHeaderMap.keys():
        #         targetFileRow.append(subHeaderMap[subHeaderKey])
        # csvWriter.writerow(targetFileRow)

        csvWriter.writerow('')

    return targetFileName

def createTargetDataFileII(targetMapList, fileNamePrefix, fileNameSuffix):

    targetFileName = 'K:/Git Code/Python/Output/'+ fileNamePrefix+'.xlsx'
    targetFileHeader = []
    targetFileRow = []
    targetFinalRowList = []
    assetSheetName = ''
    i = 0

    if os.path.isfile(targetFileName):
        workbook = load_workbook(targetFileName)
    else:
        workbook = Workbook(targetFileName)
        #defaultSheet = workbook.active
        #defaultSheet.title = 'Asset Index'
    rowValue = 1
    for targetMap in targetMapList:
        for mainHeaderKey in targetMap.keys():
            subHeaderMap = targetMap[mainHeaderKey]

            if mainHeaderKey == 'Name':
                #Need to add hyperlinks
                assetSheetName = subHeaderMap['Asset Name']
                workbook.create_sheet(assetSheetName)
                link = '#' + subHeaderMap['Asset Name']
                worksheet = workbook.get_sheet_by_name('Asset Index')
                worksheet.cell(row=rowValue,column=1).value = '=HYPERLINK("'+link+'","'+ subHeaderMap['Asset Name'] +'")'

                rowValue +=1
            if mainHeaderKey == 'Asset Details' or mainHeaderKey == 'Attributes' or mainHeaderKey == 'Relations' or mainHeaderKey == 'Complex Relations':

                for subHeaderKey in subHeaderMap.keys():

                    if '->' in subHeaderKey:
                        targetFileHeader.append(subHeaderKey[subHeaderKey.index('->') + 3:])
                    else:
                        targetFileHeader.append(subHeaderKey)
                    targetFileRow.append(subHeaderMap[subHeaderKey])

        worksheet = workbook.get_sheet_by_name(assetSheetName)
        worksheet.append(targetFileHeader)
        workbook.save(targetFileName)
        targetFinalRowList.append(targetFileRow)
        #mergeCellsInFile(2,targetFileRow,targetFileName)
        targetFileRow = []

    # for targetRow in targetFinalRowList:
    #     mergeCellsInFile(2, targetRow, targetFileName)

    return targetFileName

def createTargetDataFileIII(targetMapList, fileNamePrefix, fileNameSuffix):

    targetFileName = 'K:/Git Code/Python/Output/' + fileNamePrefix + '.xlsx'
    targetFileHeader = []
    targetFileRow = []
    targetFinalRowList = []
    columnCount = 0
    subMap = {}
    tempList = []
    mergeRow = {}
    i = 0
    if os.path.isfile(targetFileName):
        workbook = load_workbook(targetFileName)
    else:
        workbook = Workbook(targetFileRow)
        worksheet = workbook.active
        worksheet.title = 'Asset List'

        #Write the File Header
        for innerMap in targetMapList:
            for key in innerMap:
                map = innerMap[key]
                if key == 'Asset Details':
                    for innerKey in map:
                        if innerKey not in targetFileHeader:
                            targetFileHeader.append(innerKey)
                elif key == 'Attributes':
                        if 'Attribute Type' not in targetFileHeader:
                            targetFileHeader.append('Attribute Type')
                            targetFileHeader.append('Attribute')
                            columnCount += 2
                elif key == 'Relations':
                        if 'Relation Type' not in targetFileHeader:
                            targetFileHeader.append('Relation Type')
                            targetFileHeader.append('Relation')
                            columnCount += 2
                elif key == 'Complex Relations':
                        if 'Complex Relations - Relation Type' not in targetFileHeader:
                            targetFileHeader.append('Complex Relations - Relation Type')
                            targetFileHeader.append('Complex Relations - Relation')
                            targetFileHeader.append('Complex Relations - Attribute Type')
                            targetFileHeader.append('Complex Relations - Attribute')
                            columnCount += 4

        worksheet.append(targetFileHeader)

    rowNum = 2
    col = 1
    #Write actual data
    for innerMap in targetMapList:

        prevHighestLen = 0
        for key in innerMap.keys():
            map = innerMap[key]
            if key == 'Asset Details':
                for innerKey in map:
                    targetFileRow.append(map[innerKey])
                    worksheet.cell(row=rowNum, column=col).value = map[innerKey]
                    col += 1
            elif key == 'Attributes' or key == 'Relations':
                highestLen = 0
                currentRowNum = rowNum
                for innerKey in map:
                    if isinstance(map[innerKey], list):
                        highestLen += len(map[innerKey])
                    else:
                        highestLen += 1
                    targetFileRow.append(innerKey)
                    targetFileRow.append(map[innerKey])
                    worksheet.cell(row=rowNum, column=col).value = innerKey
                    col += 1
                    if isinstance(map[innerKey], list):
                        tempList = []
                        tempList = map[innerKey]
                        rowLen = len(tempList)
                        j = 0
                        for rowIndex in range(rowNum, rowNum + rowLen):
                            worksheet.cell(row=rowIndex, column=col).value = tempList[j]
                            j += 1
                        rowNum = rowNum+rowLen
                    else:
                        worksheet.cell(row=rowNum, column=col).value = map[innerKey]
                        rowNum = rowNum + 1
                    col -= 1
                rowNum = currentRowNum
                col +=2
                if prevHighestLen > highestLen:
                    highestLen = prevHighestLen
                prevHighestLen = highestLen
            elif key == 'Complex Relations':
                currentRowNum = rowNum
                for subKey in map:
                    highestLen = 0
                    subMap = map[subKey]
                    for innerKey in subMap:
                        if isinstance(subMap[innerKey], list):
                            highestLen += len(subMap[innerKey])
                        else:
                            highestLen += 1
                        targetFileRow.append(innerKey)
                        targetFileRow.append(subMap[innerKey])
                        worksheet.cell(row=rowNum, column=col).value = innerKey
                        col += 1
                        if isinstance(subMap[innerKey], list):
                            tempList = []
                            tempList = subMap[innerKey]
                            rowLen = len(tempList)
                            j = 0
                            for rowIndex in range(rowNum, rowNum + rowLen):
                                worksheet.cell(row=rowIndex, column=col).value = tempList[j]
                                j += 1
                            rowNum = rowNum + rowLen
                        else:
                            worksheet.cell(row=rowNum, column=col).value = subMap[innerKey]
                            rowNum = rowNum + 1
                        col -= 1
                    if prevHighestLen > highestLen:
                        highestLen = prevHighestLen
                    prevHighestLen = highestLen
                    rowNum = currentRowNum
                    col += 2

        targetFinalRowList.append(targetFileRow)
        mergeRow[rowNum] = highestLen
        rowNum = rowNum + highestLen + 1
        targetFileRow = []
        col = 1

    col = 1
    for rowNum in mergeRow:
         for col in range(1,5):
            worksheet.merge_cells(start_row=rowNum, start_column=col, end_row=rowNum + mergeRow[rowNum]-1,end_column=col)
            col +=1

    workbook.save(targetFileName)
    return targetFileName

if __name__ == '__main__':

    # Find the Parameter File Name
    paramFileName = sys.argv[1]

    # Open the parameter file of the business process
    os.chdir("K:\Git Code\Python\ParameterFiles")
    with open(paramFileName, 'r') as yamlFile:
        paramFileData = yaml.load(yamlFile)

    outputFilter = ''
    outputFileName = ''
    for eachKey in paramFileData.keys():
        fileNameList = []
        finalMap = {}
        finalResultList = []
        eachMap = paramFileData[eachKey]
        for itemkey in eachMap.keys():
            if itemkey == 'conditions':
                if isinstance(eachMap[itemkey], dict):
                    outputFilter = eachMap[itemkey].keys()
                    targetData = createMapII(eachMap[itemkey])
            elif itemkey == 'outputResult':
                if outputFilter != 'relationPath':
                    targetDataMap = {}
                    outputResultParameters = eachMap[itemkey]
                    for i in range(0, len(targetData)):
                        targetDataRelationsMap = {}
                        targetDataAttributesMap = {}
                        targetDataComplexRelationsMap = {}
                        targetDataComplexRelationsRelationsMap = {}
                        targetDataComplexRelationsAttributesMap = {}
                        targetDataMap = {}
                        assetName = targetData[i]['signifier']
                        for outputParameter in outputResultParameters.keys():
                            outputParameterList = outputResultParameters[outputParameter]
                            # targetDataMap['Status'] = targetData[i]['statusReference']['signifier']
                            if outputParameter == 'Asset Details':
                                for detailType in outputParameterList:
                                    targetDataMap[detailType] = fetchAssetAssetDetails(detailType, targetData[i])

                            # If needed, find the Relations, Attributes and Complex Relations of the asset
                            if outputParameter in ('Attributes', 'Relations'):
                                # Find the possible Relations and Attributes of the asset
                                possibleRelationsAndAttributesResponse = fetchPossibleRelationsAndAttributes(targetData[i]['resourceId'])
                                possibleAttributesList = []
                                possibleRelationsList = []

                                if possibleRelationsAndAttributesResponse != 'No Data Found':
                                    possibleRelationsAndAttributesList = possibleRelationsAndAttributesResponse[
                                        'representationReference']
                                    for k in range(0, len(possibleRelationsAndAttributesList)):
                                        if 'descriptionReference' in possibleRelationsAndAttributesList[k].keys():
                                            possibleAttributesList.append(possibleRelationsAndAttributesList[k]['signifier'])
                                        tempMap = {}
                                        if 'role' in possibleRelationsAndAttributesList[k].keys():
                                            tempMap['role'] = possibleRelationsAndAttributesList[k]['role']
                                            tempMap['coRole'] = possibleRelationsAndAttributesList[k]['coRole']
                                            if tempMap not in possibleRelationsList:
                                                possibleRelationsList.append(tempMap)

                                # Find the Relations
                                if outputParameter == 'Relations':
                                    if outputParameterList == 'All':
                                        targetDataRelationsMap = {}
                                        relationsResponse = fetchRelations(targetData[i]['resourceId'])
                                        if relationsResponse != 'No Data Found':
                                            # Add all the relation types that have a value
                                            for relationResponseMap in relationsResponse:
                                                for innerRelationsMap in relationResponseMap['relation']:
                                                    targetDataRelationsRolekey = innerRelationsMap['typeReference']['role']
                                                    targetDataRelationsCoRolekey = innerRelationsMap['typeReference']['coRole']
                                                    existingRoleKeyValue = []
                                                    existingCoRoleKeyValue = []
                                                    if targetDataRelationsRolekey in targetDataRelationsMap.keys():
                                                        existingRoleKeyValue = targetDataRelationsMap[targetDataRelationsRolekey]

                                                    if targetDataRelationsCoRolekey in targetDataRelationsMap.keys():
                                                        existingCoRoleKeyValue = targetDataRelationsMap[targetDataRelationsCoRolekey]

                                                    if innerRelationsMap['targetReference']['signifier'] == assetName:
                                                        existingRoleKeyValue = assetName
                                                    else:
                                                        existingRoleKeyValue.append(innerRelationsMap['targetReference']['signifier'])


                                                    if innerRelationsMap['sourceReference']['signifier'] == assetName:
                                                        existingCoRoleKeyValue = assetName
                                                    else:
                                                        existingCoRoleKeyValue.append(innerRelationsMap['sourceReference']['signifier'])

                                                    targetDataRelationsMap[innerRelationsMap['typeReference']['role']] = existingRoleKeyValue
                                                    targetDataRelationsMap[innerRelationsMap['typeReference']['coRole']] = existingCoRoleKeyValue

                                            # Add all the relation types that don't have a value
                                            for possibleRelation in possibleRelationsList:
                                                for roleMap in possibleRelation:
                                                    print(possibleRelation[roleMap])
                                                    if possibleRelation[roleMap] not in targetDataRelationsMap.keys():
                                                        targetDataRelationsMap[possibleRelation[roleMap]] = ''

                                # Find the Attributes
                                if outputParameter == 'Attributes':
                                    if outputParameterList == 'All':
                                        targetDataAttributesMap = {}
                                        attributesResponse = fetchAttributes(targetData[i]['resourceId'])
                                        attributesResponseList = attributesResponse['attributeReference']
                                        # Below block isn't required anymore.
                                        # listOfAttributesFromResponse = []
                                        # for listIndex in range(0, len(attributesResponseList)):
                                        #     attributeResponseMap = attributesResponseList[listIndex]
                                        #     listOfAttributesFromResponse.append(attributeResponseMap['labelReference']['signifier'])  # Find all the attributes of the asset that are not null
                                        # listOfAttributesFromResponse = list(set(listOfAttributesFromResponse))

                                        # Add all the attribute types that have a value
                                        for listIndex in range(0, len(attributesResponseList)):
                                            attributeResponseMap = attributesResponseList[listIndex]
                                            existingKeyValue = []
                                            # key = attributeResponseMap['resourceId'] + ' -> ' + attributeResponseMap['labelReference']['signifier']
                                            key = attributeResponseMap['labelReference']['signifier']
                                            cleanedKeyValue = cleanhtml(attributeResponseMap['value'])
                                            if key in targetDataAttributesMap.keys():
                                                existingKeyValue = targetDataAttributesMap[key]
                                            existingKeyValue.append(cleanedKeyValue)
                                            targetDataAttributesMap[key] = existingKeyValue

                                        # Add all the attribute types that don't have a value
                                        for possibleAttribute in possibleAttributesList:
                                            if possibleAttribute not in targetDataAttributesMap.keys():
                                                targetDataAttributesMap[possibleAttribute] = ''
                                    else:
                                        targetDataAttributesMap = {}
                                        attributesResponse = fetchAttributes(targetData[i]['resourceId'])
                                        listOfAttributesFromResponse = []
                                        attributesResponseList = attributesResponse['attributeReference']
                                        for listIndex in range(0, len(attributesResponseList)):
                                            attributeResponseMap = attributesResponseList[listIndex]
                                            listOfAttributesFromResponse.append(attributeResponseMap['labelReference'][
                                                                                    'signifier'])  # Find all the attributes of the asset that are not null
                                        listOfAttributesFromResponse = list(set(listOfAttributesFromResponse))

                                        # Add all the attribute types that have a value
                                        for listIndex in range(0, len(attributesResponseList)):
                                            attributeResponseMap = attributesResponseList[listIndex]
                                            existingKeyValue = []
                                            # key = attributeResponseMap['resourceId'] + ' -> ' + attributeResponseMap['labelReference']['signifier']
                                            key = attributeResponseMap['labelReference']['signifier']
                                            cleanedKeyValue = cleanhtml(attributeResponseMap['value'])
                                            if key in targetDataAttributesMap.keys():
                                                existingKeyValue = targetDataAttributesMap[key]
                                            existingKeyValue.append(cleanedKeyValue)
                                            targetDataAttributesMap[key] = existingKeyValue

                                        # Add all the attribute types that don't have a value
                                        for possibleAttribute in possibleAttributesList:
                                            if possibleAttribute not in targetDataAttributesMap.keys():
                                                targetDataAttributesMap[possibleAttribute] = ''

                            if outputParameter == 'Complex Relations':
                                # Find the Complex Relations and Attributes
                                complexRelationsMap = fetchComplexRelations(targetData[i]['resourceId'])
                                if complexRelationsMap not in ('No Relations Found', 'No Data Found'):
                                    relationValue = []
                                    for j in range(0, len(complexRelationsMap['relationReference'])):
                                        relationReferenceList = complexRelationsMap['relationReference'][j]
                                        relationKey = relationReferenceList['typeReference']['role']
                                        if relationKey in targetDataComplexRelationsRelationsMap.keys():
                                            relationValue = targetDataComplexRelationsRelationsMap[relationKey]
                                        relationValue.append(relationReferenceList['targetReference']['signifier'])
                                        targetDataComplexRelationsRelationsMap[relationKey] = relationValue
                                        relationValue = []
                                        targetDataComplexRelationsMap['Relations'] = targetDataComplexRelationsRelationsMap

                                complexRelationsAttributesMap = fetchComplexRelationsAttributes(targetData[i]['resourceId'])
                                if complexRelationsAttributesMap != 'No Attributes Found':
                                    attributeValue = []
                                    for j in range(0, len(complexRelationsAttributesMap['attributeReferences']['attributeReference'])):
                                        attributesReferenceList = complexRelationsAttributesMap['attributeReferences']['attributeReference'][j]
                                        attributeKey = attributesReferenceList['labelReference']['signifier']
                                        if attributeKey in targetDataComplexRelationsAttributesMap.keys():
                                            attributeValue = targetDataComplexRelationsAttributesMap[attributeKey]
                                        attributeValue.append(cleanhtml(attributesReferenceList['value']))
                                        targetDataComplexRelationsAttributesMap[attributeKey] = attributeValue
                                        attributeValue = []

                                    targetDataComplexRelationsMap['Attributes'] = targetDataComplexRelationsAttributesMap

                        tempMap = {}
                        finalMap['Asset Details'] = targetDataMap
                        if bool(targetDataAttributesMap):
                            finalMap['Attributes'] = targetDataAttributesMap
                        if bool(targetDataRelationsMap):
                            finalMap['Relations'] = targetDataRelationsMap
                        if bool(targetDataComplexRelationsMap):
                            finalMap['Complex Relations'] = targetDataComplexRelationsMap

                        finalResultList.append(finalMap)

                        finalMap = {}
                else:
                    print('relationPath')
            elif itemkey == 'outputFileName':
                if outputFilter != 'relationPath':
                    # Create the output file
                    fileName = createTargetDataFileIII(finalResultList, eachKey, '.csv')
                    if fileName not in fileNameList:
                       fileNameList.append(fileName)
            elif itemkey == 'email':  #Emails the output file to the audience

                print('Hello')
                # zipFilePath = "K:/Git Code/Python/Output/" + outputFileName + ".zip"
                # os.chdir("K:/Git Code/Python/Output/")
                # for file in fileNameList:
                #     with ZipFile(zipFilePath, 'a') as myzip:
                #         myzip.write(os.path.basename(file))
                #     myzip.close()
                #
                # composedMessage = MIMEMultipart()
                # composedMessage['Subject'] = 'Requested Information'
                # composedMessage['To'] = eachMap[itemkey]
                # composedMessage['From'] = "srikrishna.bingi@gmail.com"
                #
                # ctype, encoding = mimetypes.guess_type(zipFilePath)
                # if ctype is None or encoding is not None:
                #     ctype = "application/octet-stream"
                #
                # zipFile = open(zipFilePath, 'rb')
                # maintype, subtype = ctype.split("/", 1)
                # msg = MIMEBase(maintype, _subtype=subtype)
                # msg.set_payload(zipFile.read())
                # encoders.encode_base64(msg)
                #
                # msg.add_header("Content-Disposition", "attachment", filename=outputFileName+'.zip')
                # composedMessage.attach(msg)

                # mailObject = smtplib.SMTP('smtp.gmail.com', 587)
                # mailObject.starttls()
                # mailObject.login('srikrishna.bingi@gmail.com','******')
                #mailObject.sendmail("srikrishna.bingi@gmail.com","srikrishna.bingi@exusia.com",composedMessage.as_string())
                print('Done sending Mail')